from openpyxl import Workbook, load_workbook
from datetime import datetime

def update_excel(img1_name, img2_name, cosine_sim, euclidean_dist, same_person):
    file_name = "face_verification_log.xlsx"
    
    if same_person == "Verified: They are the same person.":
        try:
            workbook = load_workbook(file_name)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Date & Time", "Image 1", "Image 2", "Cosine Similarity", "Euclidean Distance", "Same Person"])

        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([current_time, img1_name, img2_name, cosine_sim, euclidean_dist, same_person])

        try:
            workbook.save(file_name)
            print(f"Results saved to {file_name}")
        except Exception as e:
            print(f"Error saving the Excel file: {e}")
